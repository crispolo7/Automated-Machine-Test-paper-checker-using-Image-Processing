import cv2
import imutils
import datetime
from imutils import contours
import numpy as np
from openpyxl import Workbook
from imutils.perspective import four_point_transform

def displayImage(img):
    cv2.imshow('Show',img)
    cv2.waitKey(0)

def rotateImage(img):
    #image = cv.imread('')
    gray = cv.cvtColor(img, cv.COLOR_BGR2GRAY)
    gray = cv.GaussianBlur(gray, (3, 3), 0)
    edged = cv.Canny(gray, 20, 100)
    cnts = cv.findContours(edged.copy(), cv.RETR_EXTERNAL, cv.CHAIN_APPROX_SIMPLE)
    cnts = cnts[0] if imutils.is_cv2() else cnts[1]
    if len(cnts) > 0:
        c = max(cnts, key=cv2.contourArea)
        mask = np.zeros(gray.shape, dtype="uint8")
        cv2.drawContours(mask, [c], -1, 255, -1)
        (x, y, w, h) = cv2.boundingRect(c)
        imageROI = image[y:y + h, x:x + w]
        maskROI = mask[y:y + h, x:x + w]
        imageROI = cv2.bitwise_and(imageROI, imageROI, mask=maskROI)
    rotated = imutils.rotate(imageROI, -90)
    rotated = imutils.rotate_bound(imageROI, -90)
    return rotated
def perspective_transform(image_name):
    img = cv.imread(image_name, 0)
    ratio = img.shape[0] / 300.0
    canny = cv.Canny(img,100,300,apertureSize = 3)
    contours, hierarchy = cv.findContours(canny,cv.RETR_LIST,cv.CHAIN_APPROX_SIMPLE)
    areas = [cv.contourArea(c) for c in contours]
    max_index = np.argmax(areas)
    cnt=contours[max_index]
    epsilon = 0.1*cv.arcLength(cnt,True)
    approx = cv.approxPolyDP(cnt,epsilon,True)
    cv.drawContours(canny, [approx], -1, (255, 255, 255), 1)
    pts = approx.reshape(4, 2)
    rect = np.zeros((4, 2), dtype = "float32")
    s = pts.sum(axis = 1)
    rect[0] = pts[np.argmin(s)]
    rect[2] = pts[np.argmax(s)]
    diff = np.diff(pts, axis = 1)
    rect[1] = pts[np.argmin(diff)]
    rect[3] = pts[np.argmax(diff)]
    rect  *  ratio
    (tl, tr, br, bl) = rect
    widthA = np.sqrt(((br[0] - bl[0]) ** 2) + ((br[1] - bl[1]) ** 2))
    widthB = np.sqrt(((tr[0] - tl[0]) ** 2) + ((tr[1] - tl[1]) ** 2))
    heightA = np.sqrt(((tr[0] - br[0]) ** 2) + ((tr[1] - br[1]) ** 2))
    heightB = np.sqrt(((tl[0] - bl[0]) ** 2) + ((tl[1] - bl[1]) ** 2))
    maxWidth = max(int(widthA), int(widthB))
    maxHeight = max(int(heightA), int(heightB))
    dst = np.array([
        [0, 0],
        [maxWidth - 1, 0],
        [maxWidth - 1, maxHeight - 1],
        [0, maxHeight - 1]], dtype = "float32")
    M = cv.getPerspectiveTransform(rect, dst)
    warp = cv.warpPerspective(img, M, (maxWidth, maxHeight))
    return warp


def gen_idnumber(thresh):
    idNumber = []
    cnts = cv2.findContours(thresh.copy(), cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    cnts = cnts[0]
    questionCnts = []

    for c in cnts:
        (x, y, w, h) = cv2.boundingRect(c)
        ar = w / float(h)

        if w >= 15 and h >= 15 and ar >= 0.9 and ar <= 1.15:
            questionCnts.append(c)

    questionCnts = contours.sort_contours(questionCnts)[0]

    for (q, i) in enumerate(np.arange(0, len(questionCnts), 5)):
        cnts = contours.sort_contours(questionCnts[i:i + 5], method="top-to-bottom")[0]
        bubbled = None

        for (j, c) in enumerate(cnts):
            mask = np.zeros(thresh.shape, dtype="uint8")
            cv2.drawContours(mask, [c], -1, 255, -1)
            mask = cv2.bitwise_and(thresh, thresh, mask=mask)
            total = cv2.countNonZero(mask)


            if bubbled is None or total > bubbled[0]:
                bubbled = (total, j)

        k = bubbled[1]
        idNumber.append(k)
    idNumber.insert(2, '-')
    id_number = ''.join(str(e) for e in idNumber)
    return id_number

def gen_key(thresh):
    answerKey = []
    cnts = cv2.findContours(thresh.copy(), cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    cnts = cnts[0]
    questionCnts = []

    for c in cnts:
        (x, y, w, h) = cv2.boundingRect(c)
        ar = w / float(h)

        if w >= 20 and h >= 20 and ar >= 0.9 and ar <= 1.1:
            questionCnts.append(c)

    questionCnts = contours.sort_contours(questionCnts, method="top-to-bottom")[0]

    for (q, i) in enumerate(np.arange(0, len(questionCnts), 5)):
        cnts = contours.sort_contours(questionCnts[i:i + 5])[0]
        bubbled = None

        for (j, c) in enumerate(cnts):
            mask = np.zeros(thresh.shape, dtype="uint8")
            cv2.drawContours(mask, [c], -1, 255, -1)
            mask = cv2.bitwise_and(thresh, thresh, mask=mask)
            total = cv2.countNonZero(mask)


            if bubbled is None or total > bubbled[0]:
                bubbled = (total, j)

        k = bubbled[1]
        k = converNtoL(k)
        answerKey.append(k)
    answerKey = ''.join(str(e) for e in answerKey)
    return answerKey

def gen_excel():
    book = Workbook()
    sheet = book.active

    sheet['A1'] = 'Image #'
    sheet['B1'] = 'ID #'
    sheet['C1'] = 'Answer'
    sheet['D1'] = 'Score'

    book.save("sample.xlsx")

def write_excel(imname, i, ans, bk):
    sheet = bk.active
    sheet.cell(row=i+1, column=1).value = imname
    sheet.cell(row=i+1, column=2).value = 'NA'
    sheet.cell(row=i+1, column=3).value = ans

def write_excel_score(s, i, bk):
    sheet = bk.active
    sheet.cell(row=i, column=4).value = s
    bk.save('sample.xlsx')





def perspective_transform2(img):

    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    blurred = cv2.GaussianBlur(gray, (5, 5), 0)
    edged = cv2.Canny(blurred, 75, 200)

    cnts = cv2.findContours(edged.copy(), cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    cnts = cnts[0] if imutils.is_cv2() else cnts[1]
    docCnt = None

    if len(cnts) > 0:
        # sort the contours according to their size in
        # descending order
        cnts = sorted(cnts, key=cv2.contourArea, reverse=True)

        # loop over the sorted contours
        for c in cnts:
            # approximate the contour
            peri = cv2.arcLength(c, True)
            approx = cv2.approxPolyDP(c, 0.02 * peri, True)

            # if our approximated contour has four points,
            # then we can assume we have found the paper
            if len(approx) == 4:
                docCnt = approx
                break

    # apply a four point perspective transform to both the
    # original image and grayscale image to obtain a top-down
    # birds eye view of the paper
    paper = four_point_transform(img, docCnt.reshape(4, 2))
    warped = four_point_transform(gray, docCnt.reshape(4, 2))

    thresh = cv2.threshold(warped, 0, 255, cv2.THRESH_BINARY_INV | cv2.THRESH_OTSU)[1]

    return thresh




def converNtoL (ans):
    if ans == 0 :
        con = 'A'

    elif ans == 1 :
        con = 'B'

    elif ans == 2 :
        con = 'C'

    elif ans == 3 :
        con = 'D'

    elif ans == 4 :
        con = 'E'

    return con


def compare(t, a):

    score = 0
    t = t.value
    a = a.value
    for j in range(0, 15):
        if a[j] == t[j]:
            score += 1
    return  score


import os


def createFolder(folderName):
    localtime = datetime.datetime.now()
    tm = str(localtime.year) + '-' + str(localtime.month)+ '-' + str(localtime.day) + '-'+ str(localtime.hour) + '-'+ str(localtime.minute)+ '-' + str(localtime.second)
    str1 = './'
    str2 = '/'
    directory = str1+folderName+tm+str2
    print directory
    try:

        if not os.path.exists(directory):
            directory = directory
            os.makedirs(directory)
            print directory
    except OSError:
        print ('Error: Creating directory. ' + directory)




# folderName = 'BSCPE5AEng'
# createFolder(folderName)
