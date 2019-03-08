from imutils.perspective import four_point_transform
from software_function import gen_key, gen_idnumber, gen_excel, perspective_transform2, write_excel, compare, write_excel_score
from imutils import contours
import numpy as np
import openpyxl
import cv2
from hardware_function import scan

def main():
    
    scan()
    gen_excel()
    papercount = count

    for i in range(1, papercount + 1):
        img = 'test_%02d.png'% i
        imgname = img
        img = cv2.imread(img)
        thresh = perspective_transform2(img)
        ans = gen_key(thresh)
        book = openpyxl.load_workbook('sample.xlsx')
        write_excel(imgname, i, ans, book)
        book.save('sample.xlsx')

    ans_keyid = 0
    ans_keyindex = 0
    book = openpyxl.load_workbook('sample.xlsx')
    sheet = book.active

    for i in range(1, papercount + 1):
        x = sheet.cell(row=i + 1, column=2)
        if x.value == '00-005':
            ans_keyid = x.value
            ans_keyindex = i

    if ans_keyid == 0 :
        print 'no answer sheet found'

    else:
        master_key = sheet.cell(row=ans_keyindex + 1, column=3)

    print master_key.value

    for i in range(1, papercount + 1):
        t_ans = sheet.cell(row=i + 1, column=3)
        score = compare(t_ans,master_key)
        write_excel_score(score, i + 1, book)

if __name__ == '__main__':
    main()





